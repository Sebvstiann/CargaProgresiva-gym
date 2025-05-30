# entrenamiento/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import RegistroEntrenamientoForm, CompletarPerfilForm
from .models import RegistroEntrenamiento, Rutina, Ejercicio
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import UsuarioRegistroForm
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from datetime import datetime

@login_required
def inicio(request):
    if not request.user.perfil_completo():
        return redirect('completar_perfil')
    return render(request, 'entrenamiento/inicio.html')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register_form'] = UsuarioRegistroForm()
        return context
    
    def post(self, request, *args, **kwargs):
        if 'register' in request.POST:
            print("Intento de registro detectado")  # Debug
            register_form = UsuarioRegistroForm(request.POST)
            print("Datos del formulario:", request.POST)  # Debug
            
            if register_form.is_valid():
                print("Formulario válido")  # Debug
                try:
                    user = register_form.save()
                    print("Usuario guardado:", user)  # Debug
                    messages.success(request, "¡Te has registrado correctamente! Ahora puedes iniciar sesión.")
                    return render(request, self.template_name, {
                        'form': self.get_form_class()(),
                        'register_form': UsuarioRegistroForm(),
                    })
                except Exception as e:
                    print("Error al guardar:", str(e))  # Debug
                    messages.error(request, f"Error al crear la cuenta: {str(e)}")
            else:
                print("Errores del formulario:", register_form.errors)  # Debug
                messages.error(request, "Por favor corrige los errores en el formulario.")
            
            return render(request, self.template_name, {
                'form': self.get_form_class()(),
                'register_form': register_form,
            })
            
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)
        if not self.request.user.perfil_completo():
            return redirect('completar_perfil')
        return response

@login_required
def completar_perfil(request):
    if request.user.perfil_completo():
        return redirect('perfil')
        
    if request.method == 'POST':
        form = CompletarPerfilForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Perfil completado correctamente!')
            return redirect('perfil')
    else:
        form = CompletarPerfilForm(instance=request.user)
    
    return render(request, 'entrenamiento/completar_perfil.html', {'form': form})

@login_required
def registrar_entrenamiento(request):
    if request.method == 'POST':
        form = RegistroEntrenamientoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.usuario = request.user
            registro.save()
            messages.success(request, '✅ Entrenamiento guardado correctamente.')
            return redirect('registrar_entrenamiento')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = RegistroEntrenamientoForm()
    return render(request, 'entrenamiento/registrar_entrenamiento.html', {'form': form})

@login_required
def exportar_registros_excel(request):
    # Obtener los filtros de la URL
    rutina_id = request.GET.get('rutina')
    ejercicio_id = request.GET.get('ejercicio')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Filtrar registros
    registros = RegistroEntrenamiento.objects.filter(usuario=request.user)
    if rutina_id:
        registros = registros.filter(rutina_id=rutina_id)
    if ejercicio_id:
        registros = registros.filter(ejercicio_id=ejercicio_id)
    if fecha_inicio:
        registros = registros.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha_hora__date__lte=fecha_fin)

    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros de Entrenamiento"

    # Escribir encabezados
    headers = [
        'Fecha',
        'Rutina',
        'Ejercicio',
        'Peso (KG)',
        'Serie 1 (reps)',
        'Serie 2 (reps)',
        'Serie 3 (reps)',
        'Rendimiento',
        'Observaciones'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Escribir datos
    for row, registro in enumerate(registros, 2):
        ws.cell(row=row, column=1, value=registro.fecha_hora.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=registro.rutina.nombre if registro.rutina else '')
        ws.cell(row=row, column=3, value=registro.ejercicio.nombre if registro.ejercicio else '')
        ws.cell(row=row, column=4, value=registro.peso_agregado_KG)
        ws.cell(row=row, column=5, value=registro.repeticiones_en_la_primera_serie)
        ws.cell(row=row, column=6, value=registro.repeticiones_en_la_segunda_serie)
        ws.cell(row=row, column=7, value=registro.repeticiones_en_la_tercera_serie)
        ws.cell(row=row, column=8, value=registro.rendimiento_percibido)
        ws.cell(row=row, column=9, value=registro.observaciones)

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=registros_entrenamiento_{}.xlsx'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )

    # Guardar el libro de Excel en la respuesta
    wb.save(response)
    return response

@login_required
def ver_registros(request):
    rutina_id = request.GET.get('rutina')
    ejercicio_id = request.GET.get('ejercicio')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    rutinas = Rutina.objects.all()
    ejercicios = Ejercicio.objects.all()

    # filtrar SOLO los registros de este usuario:
    registros = RegistroEntrenamiento.objects.filter(usuario=request.user)

    if rutina_id:
        registros = registros.filter(rutina_id=rutina_id)
    if ejercicio_id:
        registros = registros.filter(ejercicio_id=ejercicio_id)
    if fecha_inicio:
        registros = registros.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha_hora__date__lte=fecha_fin)

    return render(request, 'entrenamiento/ver_registros.html', {
        'registros': registros,
        'rutinas': rutinas,
        'ejercicios': ejercicios,
        'rutina_seleccionada': int(rutina_id) if rutina_id else None,
        'ejercicio_seleccionado': int(ejercicio_id) if ejercicio_id else None,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })

@login_required
def eliminar_registro(request, registro_id):
    registro = get_object_or_404(RegistroEntrenamiento, id=registro_id, usuario=request.user)
    if request.method == 'POST':
        registro.delete()
        messages.success(request, 'Registro eliminado correctamente.')
    return redirect('ver_registros')

@login_required
def perfil(request):
    if not request.user.perfil_completo():
        return redirect('completar_perfil')
        
    ejercicios = Ejercicio.objects.all()
    return render(request, 'entrenamiento/perfil.html', {
        'ejercicios': ejercicios
    })

@login_required
def actualizar_perfil(request):
    if request.method == 'POST':
        if 'imagen' in request.FILES:
            request.user.imagen = request.FILES['imagen']
        if request.POST.get('email'):
            request.user.email = request.POST.get('email')
        request.user.save()
        messages.success(request, 'Perfil actualizado correctamente.')
    return redirect('perfil')

@login_required
def cambiar_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        if not request.user.check_password(old_password):
            messages.error(request, 'La contraseña actual es incorrecta.')
        elif new_password1 != new_password2:
            messages.error(request, 'Las nuevas contraseñas no coinciden.')
        else:
            request.user.set_password(new_password1)
            request.user.save()
            messages.success(request, 'Contraseña actualizada correctamente.')
            return redirect('login')
    return redirect('perfil')

@login_required
def obtener_progreso(request, ejercicio_id):
    registros = RegistroEntrenamiento.objects.filter(
        usuario=request.user,
        ejercicio_id=ejercicio_id
    ).order_by('fecha_hora')

    data = {
        'fechas': [r.fecha_hora.strftime('%d/%m/%Y') for r in registros],
        'pesos': [r.peso_agregado_KG for r in registros],
        'reps1': [r.repeticiones_en_la_primera_serie for r in registros],
        'reps2': [r.repeticiones_en_la_segunda_serie for r in registros],
        'reps3': [r.repeticiones_en_la_tercera_serie for r in registros],
    }
    
    return JsonResponse(data)
